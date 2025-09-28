import psycopg2 
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
import numpy as np
from datetime import datetime

# Создаем папку для экспорта
if not os.path.exists('exports'):
    os.makedirs('exports')

def execute_complex_queries(conn):
    """Выполняет комплексные SQL-запросы для экспорта"""
    
    queries = {
        'airline_performance': """
        SELECT 
            al.airline_name as "Авиакомпания",
            al.airline_country as "Страна",
            COUNT(f.flight_id) as "Всего рейсов",
            COUNT(CASE WHEN f.status = 'On Time' THEN 1 END) as "Пунктуальные рейсы",
            ROUND(COUNT(CASE WHEN f.status = 'On Time' THEN 1 END) * 100.0 / COUNT(f.flight_id), 2) as "Пунктуальность %",
            COUNT(CASE WHEN f.status = 'Delayed' THEN 1 END) as "Задержанные рейсы",
            COUNT(CASE WHEN f.status = 'Cancelled' THEN 1 END) as "Отмененные рейсы",
            ROUND(AVG(EXTRACT(EPOCH FROM (f.scheduled_arrival - f.scheduled_departure))/3600), 2) as "Ср. продолжительность (ч)"
        FROM flights f
        JOIN airline al ON f.airline_id = al.airline_id
        GROUP BY al.airline_id, al.airline_name, al.airline_country
        HAVING COUNT(f.flight_id) > 0
        ORDER BY "Всего рейсов" DESC;
        """,
        
        'airport_traffic': """
        SELECT 
            a.airport_name as "Аэропорт",
            a.city as "Город",
            a.country as "Страна",
            COUNT(DISTINCT CASE WHEN f.departure_airport_id = a.airport_id THEN f.flight_id END) as "Рейсы на вылет",
            COUNT(DISTINCT CASE WHEN f.arrival_airport_id = a.airport_id THEN f.flight_id END) as "Рейсы на прилет",
            COUNT(DISTINCT f.flight_id) as "Общее количество рейсов",
            COUNT(DISTINCT al.airline_id) as "Количество авиакомпаний"
        FROM airport a
        LEFT JOIN flights f ON a.airport_id = f.departure_airport_id OR a.airport_id = f.arrival_airport_id
        LEFT JOIN airline al ON f.airline_id = al.airline_id
        GROUP BY a.airport_id, a.airport_name, a.city, a.country
        HAVING COUNT(DISTINCT f.flight_id) > 0
        ORDER BY "Общее количество рейсов" DESC;
        """,
        
        'passenger_activity': """
        SELECT 
            p.country_of_residence as "Страна проживания",
            COUNT(DISTINCT p.passenger_id) as "Количество пассажиров",
            COUNT(b.booking_id) as "Всего бронирований",
            COUNT(DISTINCT bf.flight_id) as "Уникальных рейсов",
            ROUND(COUNT(b.booking_id) * 1.0 / COUNT(DISTINCT p.passenger_id), 2) as "Ср. бронирований на пассажира",
            ROUND(COUNT(DISTINCT bf.flight_id) * 1.0 / COUNT(DISTINCT p.passenger_id), 2) as "Ср. рейсов на пассажира"
        FROM passengers p
        JOIN booking b ON p.passenger_id = b.passenger_id
        JOIN booking_flight bf ON b.booking_id = bf.booking_id
        GROUP BY p.country_of_residence
        HAVING COUNT(DISTINCT p.passenger_id) > 1
        ORDER BY "Всего бронирований" DESC;
        """,
        
        'monthly_statistics': """
        SELECT 
            TO_CHAR(b.created_at, 'YYYY-MM') as "Месяц",
            TO_CHAR(b.created_at, 'Month YYYY') as "Период",
            COUNT(DISTINCT b.booking_id) as "Количество бронирований",
            COUNT(DISTINCT p.passenger_id) as "Уникальные пассажиры",
            COUNT(DISTINCT bf.flight_id) as "Уникальные рейсы",
            ROUND(COUNT(DISTINCT b.booking_id) * 1.0 / COUNT(DISTINCT p.passenger_id), 2) as "Активность пассажиров"
        FROM booking b
        JOIN passengers p ON b.passenger_id = p.passenger_id
        JOIN booking_flight bf ON b.booking_id = bf.booking_id
        WHERE b.created_at IS NOT NULL
        GROUP BY "Месяц", "Период"
        ORDER BY "Месяц";
        """,
        
        'route_popularity': """
        SELECT 
            dep.airport_name as "Аэропорт вылета",
            dep.city as "Город вылета",
            arr.airport_name as "Аэропорт прилета", 
            arr.city as "Город прилета",
            COUNT(f.flight_id) as "Количество рейсов",
            COUNT(DISTINCT al.airline_id) as "Количество авиакомпаний",
            ROUND(AVG(EXTRACT(EPOCH FROM (f.scheduled_arrival - f.scheduled_departure))/3600), 2) as "Ср. время в пути (ч)"
        FROM flights f
        JOIN airport dep ON f.departure_airport_id = dep.airport_id
        JOIN airport arr ON f.arrival_airport_id = arr.airport_id
        JOIN airline al ON f.airline_id = al.airline_id
        GROUP BY dep.airport_name, dep.city, arr.airport_name, arr.city
        HAVING COUNT(f.flight_id) > 1
        ORDER BY "Количество рейсов" DESC
        LIMIT 50;
        """
    }
    
    dataframes = {}
    for name, query in queries.items():
        try:
            df = pd.read_sql_query(query, conn)
            dataframes[name] = df
            print(f"✓ Запрос '{name}': {len(df)} строк")
        except Exception as e:
            print(f"✗ Ошибка в запросе '{name}': {e}")
            # Создаем пустой DataFrame для продолжения работы
            dataframes[name] = pd.DataFrame()
    
    return dataframes

def apply_excel_formatting(writer, dataframes_dict):
    """Применяет продвинутое форматирование к Excel файлу"""
    
    # Стили для форматирования
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for sheet_name, df in dataframes_dict.items():
        if df.empty:
            continue
            
        # Записываем DataFrame в Excel
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Получаем лист
        worksheet = writer.sheets[sheet_name]
        
        # Русские названия для листов
        sheet_titles = {
            'airline_performance': 'Эффективность авиакомпаний',
            'airport_traffic': 'Трафик аэропортов', 
            'passenger_activity': 'Активность пассажиров',
            'monthly_statistics': 'Месячная статистика',
            'route_popularity': 'Популярность маршрутов'
        }
        
        worksheet.title = sheet_titles.get(sheet_name, sheet_name)
        
        # Форматируем заголовки
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = alignment
        
        # Автоматическая ширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Закрепляем первую строку и столбец
        worksheet.freeze_panes = "A2"
        
        # Добавляем фильтры на все колонки
        worksheet.auto_filter.ref = worksheet.dimensions
        
        # Условное форматирование для числовых колонок
        numeric_columns = []
        for col_idx, col_name in enumerate(df.columns, 1):
            if df[col_name].dtype in ['int64', 'float64']:
                numeric_columns.append(col_idx)
        
        for col_idx in numeric_columns:
            if col_idx > len(df.columns):
                continue
                
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            data_range = f"{col_letter}2:{col_letter}{len(df) + 1}"
            
            # Градиентная заливка (зеленый-желтый-красный)
            color_scale_rule = ColorScaleRule(
                start_type="min", start_color="FF63BE7B",  # Зеленый
                mid_type="percentile", mid_value=50, mid_color="FFFFEB84",  # Желтый
                end_type="max", end_color="FFF8696B"  # Красный
            )
            worksheet.conditional_formatting.add(data_range, color_scale_rule)
            
            # Выделение максимумов и минимумов
            if len(df) > 1:
                # Максимумы - синий
                max_rule = Rule(type="expression", formula=[f"={col_letter}2=MAX(${col_letter}$2:${col_letter}${len(df)+1})"])
                max_rule.font = Font(color="FF0000FF", bold=True)  # Синий
                worksheet.conditional_formatting.add(data_range, max_rule)
                
                # Минимумы - зеленый
                min_rule = Rule(type="expression", formula=[f"={col_letter}2=MIN(${col_letter}$2:${col_letter}${len(df)+1})"])
                min_rule.font = Font(color="FF00FF00", bold=True)  # Зеленый
                worksheet.conditional_formatting.add(data_range, min_rule)
        
        # Добавляем итоговую строку для числовых колонок
        if numeric_columns:
            total_row = len(df) + 3
            worksheet.cell(row=total_row, column=1).value = "ИТОГО:"
            worksheet.cell(row=total_row, column=1).font = Font(bold=True)
            
            for col_idx in numeric_columns:
                col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                formula = f"=SUM({col_letter}2:{col_letter}{len(df) + 1})"
                worksheet.cell(row=total_row, column=col_idx).value = formula
                worksheet.cell(row=total_row, column=col_idx).font = Font(bold=True)
        
        print(f"✓ Лист '{worksheet.title}': {len(df)} строк, {len(df.columns)} колонок")

def export_to_excel(dataframes_dict, filename):
    """
    Экспортирует словарь DataFrame в форматированный Excel файл
    
    Args:
        dataframes_dict (dict): Словарь {название_листа: DataFrame}
        filename (str): Имя файла для сохранения
    """
    
    full_path = f"exports/{filename}"
    
    try:
        with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
            # Применяем форматирование
            apply_excel_formatting(writer, dataframes_dict)
        
        # Статистика файла
        total_sheets = len(dataframes_dict)
        total_rows = sum(len(df) for df in dataframes_dict.values() if not df.empty)
        total_columns = sum(len(df.columns) for df in dataframes_dict.values() if not df.empty)
        
        print(f"\n✅ Создан файл {filename}")
        print(f"   📊 Листов: {total_sheets}")
        print(f"   📈 Строк: {total_rows}")
        print(f"   📋 Колонок: {total_columns}")
        print(f"   💾 Размер: {os.path.getsize(full_path) / 1024:.1f} KB")
        print(f"   📁 Путь: {os.path.abspath(full_path)}")
        
        return True
        
    except Exception as e:
        print(f"❌ Ошибка при создании файла {filename}: {e}")
        return False

def generate_comprehensive_report():
    """Генерирует комплексный отчет по авиаперевозкам"""
    
    print("🚀 ЗАПУСК ГЕНЕРАЦИИ КОМПЛЕКСНОГО ОТЧЕТА...")
    print("="*80)
    
    try:
        # Подключаемся к базе данных
        conn = psycopg2.connect( 
            host="localhost",
            database="airport_db",
            user="postgres",
            password="farida",  
            port="5432"
        )
        print("✓ Подключение к базе данных установлено")
        
        # Выполняем комплексные запросы
        print("\n📊 ВЫПОЛНЯЕМ SQL-ЗАПРОСЫ...")
        dataframes = execute_complex_queries(conn)
        
        # Создаем временную метку для имени файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"airport_analytics_report_{timestamp}.xlsx"
        
        # Экспортируем в Excel с форматированием
        print("\n🎨 СОЗДАЕМ ФАЙЛ EXCEL С ФОРМАТИРОВАНИЕМ...")
        success = export_to_excel(dataframes, filename)
        
        if success:
            print("\n🎉 ОТЧЕТ УСПЕШНО СОЗДАН!")
            print("="*80)
            print("📋 СОДЕРЖАНИЕ ОТЧЕТА:")
            print("   1. 📊 Эффективность авиакомпаний - KPI и метрики")
            print("   2. 🏢 Трафик аэропортов - статистика пассажиропотока") 
            print("   3. 👥 Активность пассажиров - поведенческая аналитика")
            print("   4. 📅 Месячная статистика - временные тренды")
            print("   5. 🛫 Популярность маршрутов - топ направлений")
            print("\n💡 ФУНКЦИОНАЛ ФАЙЛА:")
            print("   • 🔒 Закрепленные заголовки")
            print("   • 🎨 Градиентное форматирование числовых данных")
            print("   • 🔍 Фильтры по всем колонкам")
            print("   • 📈 Выделение минимумов/максимумов")
            print("   • 📊 Автоматические итоги")
            print("   • 📏 Оптимальная ширина колонок")
            
        return success
        
    except Exception as e:
        print(f"❌ Критическая ошибка: {e}")
        return False
    finally:
        if 'conn' in locals():
            conn.close()
            print("\n🔒 Подключение к базе данных закрыто")

# Дополнительная функция для быстрого экспорта отдельных DataFrame
def quick_export_single_df(df, sheet_name, filename_prefix="quick_export"):
    """Быстрый экспорт одного DataFrame с базовым форматированием"""
    
    if df.empty:
        print("❌ DataFrame пустой, экспорт невозможен")
        return False
    
    timestamp = datetime.now().strftime("%H%M%S")
    filename = f"exports/{filename_prefix}_{timestamp}.xlsx"
    
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            
            # Базовое форматирование
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            
        print(f"✅ Быстрый экспорт: {filename} ({len(df)} строк)")
        return True
        
    except Exception as e:
        print(f"❌ Ошибка быстрого экспорта: {e}")
        return False

if __name__ == "__main__":
    # Генерируем комплексный отчет
    generate_comprehensive_report()
    
